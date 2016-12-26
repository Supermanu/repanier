# -*- coding: utf-8
from __future__ import unicode_literals

from django.http import HttpResponse
from django.utils import translation
from django.utils.text import slugify
from django.utils.translation import ugettext_lazy as _
from openpyxl.style import Border
from openpyxl.style import NumberFormat

import repanier.apps
from export_tools import *
from repanier.models import Configuration
from repanier.const import *
from repanier.models import BankAccount, Purchase
from repanier.models import Customer
from repanier.models import CustomerInvoice
from repanier.models import Producer
from repanier.models import ProducerInvoice
from repanier.tools import get_invoice_unit
from repanier.xlsx.xlsx_stock import export_permanence_stock


def export_invoice(permanence, customer=None, producer=None, wb=None, sheet_name=EMPTY_STRING):
    ws = None
    # Detail of what has been prepared
    if customer is None and producer is None:
        wb, ws = new_landscape_a4_sheet(wb, "%s %s" % (_('Account summary'), sheet_name),
                                        permanence)

        row_num = 0

        bank_account = BankAccount.objects.filter(
            permanence_id=permanence.id,
            customer__isnull=True,
            producer__isnull=True
        ).order_by('?').first()
        if bank_account is None:
            # Permanence not invoiced yet : Nothing to do
            return wb
        customer_set = Customer.objects.filter(customerinvoice__isnull=False).distinct()
        for customer in customer_set:
            bank_amount_in = bank_amount_out = DECIMAL_ZERO
            prepared = DECIMAL_ZERO
            customer_invoice = CustomerInvoice.objects.filter(
                customer_id=customer.id,
                permanence_id=permanence.id
            ).order_by('?').first()
            if customer_invoice is not None:
                balance_before = customer_invoice.previous_balance.amount
                bank_amount_in = customer_invoice.bank_amount_in.amount
                bank_amount_out = customer_invoice.bank_amount_out.amount
                if customer_invoice.customer_id == customer_invoice.customer_who_pays_id:
                    prepared = customer_invoice.get_total_price_with_tax().amount
                balance_after = customer_invoice.balance.amount
            else:
                last_customer_invoice = CustomerInvoice.objects.filter(
                    customer_id=customer.id,
                    invoice_sort_order__lte=bank_account.id
                ).order_by('-invoice_sort_order').first()
                if last_customer_invoice is not None:
                    balance_before = last_customer_invoice.balance.amount
                    balance_after = last_customer_invoice.balance.amount
                else:
                    # No invoice yet.
                    balance_before = customer.initial_balance.amount
                    balance_after = customer.initial_balance.amount

            row = [
                (_('Name'), 40, customer.long_basket_name, NumberFormat.FORMAT_TEXT),
                (_('Balance before'), 15, balance_before, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('bank_amount_in'), 10, bank_amount_in, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('bank_amount_out'), 10, bank_amount_out, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Prepared'), 10, prepared, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Balance after'), 15, balance_after, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Name'), 20, customer.short_basket_name, NumberFormat.FORMAT_TEXT),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in range(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = row[col_num][ROW_VALUE]
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]

            row_num += 1

        customer = None
        row_break = row_num
        row_num += 1

        producer_set = Producer.objects.filter(producerinvoice__isnull=False).distinct()
        for producer in producer_set:
            bank_amount_in = bank_amount_out = payment = DECIMAL_ZERO
            prepared = DECIMAL_ZERO
            producer_invoice = ProducerInvoice.objects.filter(
                producer_id=producer.id,
                permanence_id=permanence.id,
            ).order_by('?').first()
            if producer_invoice is not None:
                balance_before = -producer_invoice.previous_balance.amount
                bank_amount_in = producer_invoice.bank_amount_in.amount
                bank_amount_out = producer_invoice.bank_amount_out.amount
                prepared = producer_invoice.get_total_price_with_tax().amount
                balance_after = -producer_invoice.balance.amount
            else:
                last_producer_invoice = ProducerInvoice.objects.filter(
                    producer_id=producer.id,
                    invoice_sort_order__lte=bank_account.id
                ).order_by('-invoice_sort_order').first()
                if last_producer_invoice is not None:
                    balance_before = -last_producer_invoice.balance.amount
                    balance_after = -last_producer_invoice.balance.amount
                else:
                    # No invoice yet.
                    balance_before = -producer.initial_balance.amount
                    balance_after = -producer.initial_balance.amount

            row = [
                (_('Name'), 40, producer.long_profile_name, NumberFormat.FORMAT_TEXT),
                (_('Balance before'), 15, balance_before, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('bank_amount_in'), 10, bank_amount_in, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('bank_amount_out'), 10, bank_amount_out, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Prepared'), 10, prepared, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Balance after'), 15, balance_after, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
                (_('Name'), 20, producer.short_profile_name, NumberFormat.FORMAT_TEXT),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in range(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = row[col_num][ROW_VALUE]
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]

            row_num += 1

        producer = None
        final_bank_amount = bank_account.bank_amount_in.amount - bank_account.bank_amount_out.amount
        bank_account = BankAccount.objects.filter(
            id__lt=bank_account.id,
            customer__isnull=True,
            producer__isnull=True
        ).order_by("-id").first()
        if bank_account is not None:
            initial_bank_amount = bank_account.bank_amount_in.amount - bank_account.bank_amount_out.amount
        else:
            # This shouldn't occur because an initial balance is automatically generated
            # if not present
            # when invoicing the very first permanence.
            initial_bank_amount = DECIMAL_ZERO
        row_num += 1
        c = ws.cell(row=row_num, column=1)
        c.value = initial_bank_amount
        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
        c = ws.cell(row=row_num, column=4)
        formula = 'B%s+SUM(C%s:C%s)-SUM(D%s:D%s)' % (row_num + 1, 2, row_num - 1, 2, row_num - 1)
        c.value = '=' + formula
        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

        row_num += 1
        c = ws.cell(row=row_num, column=4)
        formula = 'SUM(F%s:F%s)-SUM(F%s:F%s)' % (2, row_break, row_break + 2, row_num - 2)
        c.value = '=' + formula
        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

        row_num += 1
        c = ws.cell(row=row_num, column=4)
        c.value = final_bank_amount
        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

        ws = None

        purchase_set = Purchase.objects.filter(
            permanence_id=permanence.id,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__preparation_sort_order",
            "customer__short_basket_name"
        ).distinct()
        hide_producer_prices = False
        hide_customer_prices = False
    elif customer is not None:
        purchase_set = Purchase.objects.filter(
            permanence_id=permanence.id, customer=customer,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__preparation_sort_order",
        ).distinct()
        hide_producer_prices = True
        hide_customer_prices = False
    else:
        purchase_set = Purchase.objects.filter(
            permanence_id=permanence.id, producer=producer,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__preparation_sort_order",
        ).distinct()
        hide_producer_prices = False
        hide_customer_prices = True

    row_num = 0

    hide_column_deposit = True

    for purchase in purchase_set:

        if ws is None:
            if producer is not None:
                # To the producer we speak of "payment".
                # This is the detail of the payment to the producer, i.e. received products
                wb, ws = new_landscape_a4_sheet(wb, "%s %s" % (_('Payment'), sheet_name),
                                                permanence)
            else:
                # To the customer we speak of "invoice".
                # This is the detail of the invoice, i.e. sold products
                wb, ws = new_landscape_a4_sheet(wb, "%s %s" % (_('Invoice'), sheet_name),
                                                permanence)

        qty = purchase.quantity_invoiced

        if purchase.offer_item.unit_deposit.amount != DECIMAL_ZERO:
            hide_column_deposit = False

        unit = get_invoice_unit(order_unit=purchase.offer_item.order_unit, qty=qty)
        row = [
            (_("Producer"), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
            (_("Basket"), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
            (_("Department"), 15,
             purchase.offer_item.department_for_customer.short_name if purchase.offer_item.department_for_customer is not None else EMPTY_STRING,
             NumberFormat.FORMAT_TEXT, False),
            (_("Product"), 60, purchase.get_long_name(), NumberFormat.FORMAT_TEXT, False),
            (_("Quantity"), 10, qty, '#,##0.????',
             True if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG else False),
            (_("Unit"), 10, unit, NumberFormat.FORMAT_TEXT, False),
            (_("deposit"), 10, purchase.offer_item.unit_deposit.amount, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX,
             False)]
        if hide_producer_prices:
            row += [
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
            ]
        else:
            row += [
                (_("producer unit price"), 10, purchase.get_producer_unit_price(),
                 repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                (_("purchase price"), 10, purchase.purchase_price.amount, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX,
                 False),
                (_("Vat"), 10,
                 (purchase.offer_item.producer_vat.amount * purchase.get_quantity()).quantize(FOUR_DECIMALS),
                 repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False)
            ]

        if hide_customer_prices:
            row += [
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
            ]
        else:
            row += [
                (_("customer unit price"), 10, purchase.get_customer_unit_price(),
                 repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                (_("selling price"), 10, purchase.selling_price.amount,
                 repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                (_("Vat"), 10,
                (purchase.offer_item.customer_vat.amount * purchase.get_quantity()).quantize(FOUR_DECIMALS),
                 repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
            ]
        if hide_producer_prices and hide_customer_prices:
            row += [
                (EMPTY_STRING, 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
            ]
        else:
            row += [
                (EMPTY_STRING, 10, purchase.get_vat_level_display(), NumberFormat.FORMAT_TEXT, False)
            ]
        row += [
            (_("comment"), 30, EMPTY_STRING if purchase.comment is None else purchase.comment, NumberFormat.FORMAT_TEXT,
             False),
        ]

        if row_num == 0:
            worksheet_set_header(ws, row)
            row_num += 1

        for col_num in range(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = "%s" % (row[col_num][ROW_VALUE])
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]
            if row[col_num][ROW_BOX]:
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.borders.left.border_style = Border.BORDER_THIN
                c.style.borders.right.border_style = Border.BORDER_THIN
            else:
                c.style.borders.bottom.border_style = Border.BORDER_HAIR
            if col_num == 7:
                c.style.font.bold = True

        row_num += 1

    if wb is not None and ws is not None:
        if hide_column_deposit:
            ws.column_dimensions[get_column_letter(7)].visible = False
        if hide_producer_prices:
            ws.column_dimensions[get_column_letter(8)].visible = False
            ws.column_dimensions[get_column_letter(9)].visible = False
            ws.column_dimensions[get_column_letter(10)].visible = False
        if hide_customer_prices:
            ws.column_dimensions[get_column_letter(11)].visible = False
            ws.column_dimensions[get_column_letter(12)].visible = False
            ws.column_dimensions[get_column_letter(13)].visible = False
        if hide_producer_prices and hide_customer_prices:
            ws.column_dimensions[get_column_letter(14)].visible = False
        if customer is not None or producer is not None:
            for col_num in range(14):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 1:
                    c.value = _("Total Price") + " " + sheet_name
                    c.style.font.bold = True
                if col_num == 8:
                    formula = 'SUM(I%s:I%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 9:
                    formula = 'SUM(J%s:J%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 11:
                    formula = 'SUM(L%s:L%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 12:
                    formula = 'SUM(M%s:M%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
        if customer is not None:
            config = Configuration.objects.get(id=DECIMAL_ONE)
            group_label = config.group_label
            if group_label:
                row_num +=1
                for col_num in range(14):
                    c = ws.cell(row=row_num, column=col_num)
                    c.style.borders.top.border_style = Border.BORDER_THIN
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    if col_num == 1:
                        c.value = group_label
                        c.style.font.bold = True
    return wb


def admin_export(request, queryset):
    permanence = queryset.first()
    wb = export_invoice(permanence=permanence, sheet_name=repanier.apps.REPANIER_SETTINGS_GROUP_NAME)
    wb = export_permanence_stock(permanence=permanence, wb=wb, ws_customer_title=None)
    if wb is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = "attachment; filename={0}-{1}.xlsx".format(
            slugify(_("Accounting report")),
            slugify(permanence)
        )
        wb.save(response)
        return response
    else:
        return None